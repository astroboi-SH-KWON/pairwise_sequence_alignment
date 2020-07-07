import glob
import openpyxl

import Logic

class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_dat = ".dat"
        self.ext_xlsx = ".xlsx"

    """
    get file lists in target dir by target ext
    :param
        path : target dir + "*." + target ext
    :return
        ['target dir/file_name.target ext', 'target dir/file_name.target ext' ...]
    """
    def get_files_from_dir(self, path):
        return glob.glob(path)

    """
    :param
        result_dict = { 'file_name as key' : 
                            [['Final index', sub_dict, ins_dict, del_dict, last index of ref_seq], ...]
                        , 'D:/000_WORK/YuGooSang_KimHuiKwon/20200609/WORK_DIR/first_excel_output\\result_gDNA_0609.txt': 
                            [['Group1,2_RT/20-PBS/7-#Target723'
                                    , {1: 'T->A', 36: 'T->C', 88: 'T->A', 137: 'A->G'}
                                    , {35: 'A', 38: 'GGG', 49: 'GGGGT'}
                                    , {31: 'C', 32: 'C', 42: 'A', 43: 'A', 44: 'G', 45: 'A', 46: 'A', 47: 'A', 48: 'A'}, 284]
                            , ['Group1,2_RT/12-PBS/11-#Target1948', {}, {}, {}, 280]
                            , ['Group1,2_RT/20-PBS/17-#Target833'
                                    , {}
                                    , {294: 'CGTGCACGTGACACGTTCCAGACCGTACATGCTTACATGGGATGAAGCTTGGCGTAACTAGATCTTGAGACAAATGGCAGTATT'}
                                    , {1: 'T', 2: 'A', 3: 'T', 4: 'A', 5: 'T', 6: 'A', 7: 'T', 8: 'C', 9: 'T', 10: 'T', 11: 'G', 12: 'T', 13: 'G', 14: 'G', 15: 'A', 16: 'A', 17: 'A', 18: 'G', 19: 'G', 20: 'A', 21: 'C', 22: 'G', 23: 'A', 24: 'A', 25: 'A', 26: 'C', 27: 'A', 28: 'C', 29: 'C', 30: 'G', 31: 'C', 32: 'A', 33: 'T', 34: 'C', 35: 'T', 36: 'G', 37: 'G', 38: 'A', 39: 'G', 40: 'G', 41: 'A', 42: 'T', 43: 'T', 44: 'G', 45: 'T', 46: 'G', 47: 'A', 48: 'T', 49: 'C', 50: 'G', 51: 'T', 52: 'T', 53: 'T', 54: 'T', 55: 'A', 56: 'G', 57: 'A', 58: 'G', 59: 'C', 60: 'T', 61: 'A', 62: 'G', 63: 'A', 64: 'A', 65: 'A', 66: 'T', 67: 'A', 68: 'G', 69: 'C', 70: 'A', 71: 'A', 72: 'G', 73: 'T', 74: 'T', 75: 'A', 76: 'A', 77: 'A', 78: 'A', 79: 'T', 80: 'A', 81: 'A', 82: 'G', 83: 'G', 84: 'C', 85: 'T', 86: 'A', 87: 'G', 88: 'T', 89: 'C', 92: 'T', 93: 'T', 94: 'A', 95: 'T', 96: 'C', 97: 'A', 98: 'A'}
                                    , 294]
                            , ['Group1,2_RT/12-PBS/9-#Target489'
                                    , {33: 'T->C', 34: 'T->G', 35: 'C->G', 37: 'C->A', 49: 'A->T', 222: 'C->T'}
                                    , {29: 'G', 40: 'GT'}
                                    , {43: 'G', 44: 'C', 48: 'A'}
                                    , 278]
                            ]
                    }
    """

    def make_excel(self, path, result_dict):
        logic = Logic.Logics()

        for fn_key, val_list in result_dict.items():
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            row = 1
            sheet.cell(row=row, column=1, value="index")
            sheet.cell(row=row, column=2, value='Final index')
            sheet.cell(row=row, column=3, value='Type')
            sheet.cell(row=row, column=4, value='Start')
            sheet.cell(row=row, column=5, value='End')
            sheet.cell(row=row, column=6, value='Sequence')

            for val_arr in val_list:
                final_index = val_arr[0].replace('"', '')
                sub_dict = val_arr[1]
                ins_dict = val_arr[2]
                del_dict = val_arr[3]
                last_idx = val_arr[4]
                read_no = val_arr[5]
                end_idx = 0
                for i in range(last_idx):
                    if i >= end_idx:
                        if i in sub_dict:
                            row += 1
                            # sheet.cell(row=row, column=1, value=str(row - 1))
                            sheet.cell(row=row, column=1, value=str(read_no))
                            sheet.cell(row=row, column=2, value=final_index)
                            sheet.cell(row=row, column=3, value='Sub')
                            sheet.cell(row=row, column=4, value=str(i))
                            end_idx, sub_seq_from, sub_seq_to = logic.get_sub_idx_seq(i + 1, sub_dict,
                                                                                      sub_dict[i].split("->"))
                            sheet.cell(row=row, column=5, value=str(end_idx))
                            sheet.cell(row=row, column=6, value=sub_seq_from + "->" + sub_seq_to)

                        elif i in del_dict:
                            row += 1
                            # sheet.cell(row=row, column=1, value=str(row - 1))
                            sheet.cell(row=row, column=1, value=str(read_no))
                            sheet.cell(row=row, column=2, value=final_index)
                            sheet.cell(row=row, column=3, value='Del')
                            sheet.cell(row=row, column=4, value=str(i))
                            end_idx, del_seq = logic.get_del_idx_seq(i + 1, del_dict, del_dict[i])
                            sheet.cell(row=row, column=5, value=str(end_idx))
                            sheet.cell(row=row, column=6, value=del_seq)

                        elif i in ins_dict:
                            row += 1
                            # sheet.cell(row=row, column=1, value=str(row - 1))
                            sheet.cell(row=row, column=1, value=str(read_no))
                            sheet.cell(row=row, column=2, value=final_index)
                            sheet.cell(row=row, column=3, value='Ins')
                            if i == 0:
                                sheet.cell(row=row, column=4, value=str(1))
                            else:
                                sheet.cell(row=row, column=4, value=str(i))
                            if i == last_idx:
                                sheet.cell(row=row, column=5, value=str(i))
                            else:
                                sheet.cell(row=row, column=5, value=str(i + 1))
                            sheet.cell(row=row, column=6, value=ins_dict[i])

            workbook.save(filename=path + fn_key[fn_key.index("input") + len("input") + 1:].replace(".txt", "_") + self.ext_xlsx)

    def read_tb_txt(self, path):
        result_list = []
        with open(path, "r") as f:
            f.readline().replace("\n", "")
            while True:
                tmp_line = f.readline().replace("\n", "")
                if tmp_line == "":
                    break

                tmp_arr = tmp_line.split("\t")
                result_list.append(tmp_arr)

        return result_list

    def make_excel_mutil_processing(self, path, result_dict):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="index")
        sheet.cell(row=row, column=2, value='Ref Sequence')
        sheet.cell(row=row, column=3, value='needle_cnt')
        sheet.cell(row=row, column=4, value='ins')
        sheet.cell(row=row, column=5, value='del')
        sheet.cell(row=row, column=6, value='sub')
        sheet.cell(row=row, column=7, value='NGS read')

        for ref_seq_key, val_list in result_dict.items():
            for val_arr in val_list:
                row += 1
                sheet.cell(row=row, column=1, value=str(row - 1))
                sheet.cell(row=row, column=2, value=ref_seq_key)
                sheet.cell(row=row, column=3, value=str(val_arr[0]))
                sheet.cell(row=row, column=4, value=str(val_arr[1]))
                sheet.cell(row=row, column=5, value=str(val_arr[2]))
                sheet.cell(row=row, column=6, value=str(val_arr[3]))
                sheet.cell(row=row, column=7, value=val_arr[4])

        workbook.save(
            filename=path + self.ext_xlsx)

    def make_txt_needle_result(self, path, result_dict, path_split):
        for fn_key, val_list in result_dict.items():
            fn = fn_key[fn_key.index(path_split) + len(path_split) + 1:].replace(self.ext_txt, "")
            with open(path + fn + self.ext_txt, "a") as f:
                for val_arr in val_list:
                    f.write("Read No. " + val_arr[0] + "\n")
                    f.write(val_arr[1] + "\n")
                    f.write(val_arr[2])
                    f.write("\n")

    def read_tb_txt_wo_header(self, path):
        result_list = []
        with open(path, "r") as f:
            idx = 0
            while True:
                tmp_line = f.readline().replace("\n", "")
                if tmp_line == "":
                    break
                idx += 1
                tmp_arr = tmp_line.split("\t")
                tmp_arr.append(idx)
                result_list.append(tmp_arr)

        return result_list

    def make_excel_simple(self, path, result_dict):
        logic = Logic.Logics()

        for fn_key, val_list in result_dict.items():
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            row = 1
            sheet.cell(row=row, column=1, value="ngs_id")
            sheet.cell(row=row, column=2, value='Type')
            sheet.cell(row=row, column=3, value='Start')
            sheet.cell(row=row, column=4, value='End')
            sheet.cell(row=row, column=5, value='Sequence')

            for val_arr in val_list:
                sub_dict = val_arr[0]
                ins_dict = val_arr[1]
                del_dict = val_arr[2]
                last_idx = val_arr[3]
                ngs_id = val_arr[4]

                end_idx = 0
                for i in range(last_idx + 1):
                    if i >= end_idx:
                        if i in sub_dict:
                            row += 1
                            sheet.cell(row=row, column=1, value=str(ngs_id))
                            sheet.cell(row=row, column=2, value='Sub')
                            sheet.cell(row=row, column=3, value=str(i))
                            end_idx, sub_seq_from, sub_seq_to = logic.get_sub_idx_seq(i + 1, sub_dict,
                                                                                      sub_dict[i].split("->"))
                            sheet.cell(row=row, column=4, value=str(end_idx))
                            sheet.cell(row=row, column=5, value=sub_seq_from + "->" + sub_seq_to)

                        elif i in del_dict:
                            row += 1
                            sheet.cell(row=row, column=1, value=str(ngs_id))
                            sheet.cell(row=row, column=2, value='Del')
                            sheet.cell(row=row, column=3, value=str(i))
                            end_idx, del_seq = logic.get_del_idx_seq(i + 1, del_dict, del_dict[i])
                            sheet.cell(row=row, column=4, value=str(end_idx))
                            sheet.cell(row=row, column=5, value=del_seq)

                        elif i in ins_dict:
                            row += 1
                            sheet.cell(row=row, column=1, value=str(ngs_id))
                            sheet.cell(row=row, column=2, value='Ins')
                            if i == 0:
                                sheet.cell(row=row, column=3, value=str(1))
                            else:
                                sheet.cell(row=row, column=3, value=str(i))
                            if i == last_idx:
                                sheet.cell(row=row, column=4, value=str(i))
                            else:
                                sheet.cell(row=row, column=4, value=str(i + 1))
                            sheet.cell(row=row, column=5, value=ins_dict[i])

            workbook.save(filename=path + self.ext_xlsx)
